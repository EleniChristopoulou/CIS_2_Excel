import fitz
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# === GUI to select PDF ===
def select_pdf():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select PDF file",
        filetypes=[("PDF Files", "*.pdf")]
    )

# === Structured section extractor function ===
def extract_structured_section(pdf_path, search_text: str):
    stop_phrase = "Default Value"
    pattern = re.compile(re.escape(search_text).replace(r"\ ", r"\s+"), re.IGNORECASE)

    section_pattern = re.compile(
        r"Profile Applicability:\s*(.*?)\s*"
        r"Description:\s*(.*?)\s*"
        r"Rationale:\s*(.*?)\s*"
        r"Impact:\s*(.*?)\s*"
        r"Audit:\s*(.*?)\s*"
        r"Remediation:\s*(.*?)\s*(?=Default Value|$)",
        re.IGNORECASE | re.DOTALL
    )

    doc = fitz.open(pdf_path)
    match_count = 0
    collecting_text = ""

    for page_number, page in enumerate(doc):
        page_text = page.get_text("text")
        text_norm = re.sub(r"\s+", " ", page_text)

        if not collecting_text:
            match = pattern.search(text_norm)
            if match:
                match_count += 1
                if match_count == 2:
                    collecting_text += text_norm[match.end():] + " "
        else:
            collecting_text += text_norm + " "

        section_match = section_pattern.search(collecting_text)
        if section_match:
            doc.close()
            return [group.strip() for group in section_match.groups()]

    doc.close()
    return [None]*6  # if not found

# === Extract controls and write to Excel ===
def extract_controls_with_parent(pdf_path):
    doc = fitz.open(pdf_path)
    toc = doc.get_toc()
    doc.close()

    if not toc:
        print("No embedded Table of Contents found.")
        return

    inside = False
    hierarchy = {}
    rows = []

    print("\n===== AUTOMATED / MANUAL CONTROLS =====\n")

    for level, title, page in toc:
        clean_title = title.strip()

        # Start at Recommendations
        if clean_title.startswith("Recommendations"):
            inside = True

        # Stop at Appendix
        if clean_title.startswith("Appendix"):
            break

        if not inside:
            continue

        hierarchy[level] = clean_title

        # Only keep controls with (Manual) or (Automated)
        if "(Manual)" in clean_title or "(Automated)" in clean_title:
            parent_level = level - 1
            parent_title = hierarchy.get(parent_level, "")

            # Extract section number and name
            section_match = re.match(r"^([\d\.]+)\s+(.*)", parent_title)
            if section_match:
                section_number = section_match.group(1)
                section_name = section_match.group(2)
            else:
                section_number = ""
                section_name = parent_title

            # Extract control number and description
            control_match = re.match(r"^([\d\.]+)\s+(.*)", clean_title)
            if control_match:
                control_number = control_match.group(1)
                control_desc = control_match.group(2)
            else:
                control_number = ""
                control_desc = clean_title

            # Extract structured section fields
            profile_applicability, description, rationale, impact, audit, remediation = extract_structured_section(pdf_path, clean_title)

            print(f"{section_number} | {section_name} | {control_number} | {control_desc}")

            # Append row with all info
            rows.append({
                "Section's Number": section_number,
                "Section": section_name,
                "Control's Number": control_number,
                "Control": control_desc,
                "Profile Applicability": profile_applicability,
                "Description": description,
                "Rationale": rationale,
                "Impact": impact,
                "Audit": audit,
                "Remediation": remediation
            })

    if rows:
        # Save to Excel first
        df = pd.DataFrame(rows)
        excel_path = pdf_path.replace(".pdf", "_controls.xlsx")
        df.to_excel(excel_path, index=False)

        # === Apply styling with openpyxl ===
        wb = load_workbook(excel_path)
        ws = wb.active

        # Style headers: bold, bigger font, black
        header_font = Font(bold=True, size=12, color="000000")
        for cell in ws[1]:
            cell.font = header_font

        # Row colors based on Profile Applicability
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            profile_cell = row[4]  # 5th column: Profile Applicability
            value = profile_cell.value
            if value:
                value = value.strip().upper()
                if "BL" in value:
                    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # light grey
                elif "L1" in value:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
                elif "L2" in value:
                    fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
                else:
                    fill = None

                if fill:
                    for cell in row:
                        cell.fill = fill

        # Save styled Excel
        styled_path = excel_path.replace(".xlsx", "_controls.xlsx")
        wb.save(styled_path)
        print(f"\nStyled Excel saved as: {styled_path}")

if __name__ == "__main__":
    pdf_file = select_pdf()
    if pdf_file:
        extract_controls_with_parent(pdf_file)
    else:
        print("No file selected.")
